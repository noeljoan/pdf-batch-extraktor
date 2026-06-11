"""
===========================================================================
PROJEKT: Universal KI Normen-Batch-Extraktor (Local LLM)
AUTOR: Noel Joan
VERSION: 14.0 (Final)
LIZENZ: Open Source / Freie Verwendung
===========================================================================
Dieses Tool verarbeitet pdf-Dateien komplett
lokal auf dem PC des Nutzers. Es nutzt PyMuPDF, Tesseract-OCR und die
ressourcenschonende lokale KI 'Qwen2.5:1.5b' via Ollama.
"""

import os
import sys
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import ollama  # Lokale KI-Schnittstelle

# ---------------------------------------------------------------------------
# Tesseract Pfad-Konfiguration
# ---------------------------------------------------------------------------
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

# ---------------------------------------------------------------------------
# KI-gestützte Bereinigungs- und Extraktions-Logik
# ---------------------------------------------------------------------------

def remove_illegal_excel_characters(text: str) -> str:
    """Entfernt alle illegalen ASCII- und Unicode-Steuerzeichen für Excel."""
    if not text:
        return ""
    illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
    clean_text = illegal_chars.sub("", text)
    return "".join(c for c in clean_text if c.isprintable() or c in ['\n', '\r', '\t'])

def ask_local_ai_to_clean(ocr_text: str) -> dict:
    """Nutzt das schnelle qwen2.5:1.5b Modell mit optimiertem Titel-Endungsfilter."""
    parsed_data = {"Normnummer": "Nicht gefunden", "Ausgabedatum": "Nicht gefunden", "Titel_DE": "Nicht gefunden"}
    
    prompt = f"""
    Du bist ein Experte für Dokumentenanalyse. Analysiere diesen fehlerhaften OCR-Text einer pdf-Datei.
    Extrahiere die drei Felder exakt im vorgegebenen Format. Wenn ein Feld nicht existiert oder unleserlich ist, schreibe "Nicht gefunden".
    Entferne jegliche Verlagsdaten, Copyright-Hinweise, Adressen, Telefonnummern oder reinen Zeichensalat.
    
    WICHTIGE KOSMETISCHE REGEL:
    Kürze oder bereinige den Titel am Ende so, dass er NIEMALS mit einzelnen, unvollständigen Wörtern oder einsamen Anhängseln wie 
    "Anhang", "Teil", "Norm", "Blatt", "Klasse" oder einzelnen Buchstaben endet, wenn dadurch der Fluss gestört wird. 

    Format für die Rückgabe (Gib AUSSCHLIESSLICH diese drei Zeilen zurück! Keine Einleitung, keine Erklärung!):
    NORM: [Hier die bereinigte Normnummer, z.B. DIN 13]
    DATUM: [Hier das Ausgabedatum, z.B. August 1968]
    TITEL: [Hier der saubere, offizielle Titel des Dokuments]

    OCR-Text:
    \"\"\"{ocr_text}\"\"\"
    """
    
    try:
        response = ollama.generate(model='qwen2.5:1.5b', prompt=prompt, options={"temperature": 0.0})
        response_text = response.get('response', '')
        
        norm_match = re.search(r'NORM:\s*(.*)', response_text, re.IGNORECASE)
        date_match = re.search(r'DATUM:\s*(.*)', response_text, re.IGNORECASE)
        titel_match = re.search(r'TITEL:\s*(.*)', response_text, re.IGNORECASE)
        
        if norm_match and "nicht gefunden" not in norm_match.group(1).lower():
            parsed_data["Normnummer"] = norm_match.group(1).strip()
        if date_match and "nicht gefunden" not in date_match.group(1).lower():
            parsed_data["Ausgabedatum"] = date_match.group(1).strip()
        if titel_match and "nicht gefunden" not in titel_match.group(1).lower():
            parsed_data["Titel_DE"] = titel_match.group(1).strip()
            
    except Exception as e:
        print(f"KI-Fehler: {e}")
        
    return parsed_data

def extract_metadata_from_pdf_content(pdf_path: str) -> dict:
    """Rendert die 1. Seite hocheffizient als Bild für das lokale KI-Auge."""
    filename = os.path.basename(pdf_path)
    
    data = {
        "Dateiname": filename,
        "Normnummer": "Nicht gefunden",
        "Ausgabedatum": "Nicht gefunden",
        "Titel_DE": "Nicht gefunden",
        "Status": "Unbekannt"
    }
    
    if not os.path.exists(TESSERACT_PATH):
        raise FileNotFoundError(f"Tesseract-OCR wurde unter {TESSERACT_PATH} nicht gefunden!")
    
    try:
        doc = fitz.open(pdf_path)
        if doc.is_encrypted:
            doc.authenticate("")
            
        if len(doc) > 0:
            page = doc[0]  # Erste Seite greifen
            pix = page.get_pixmap(matrix=fitz.Matrix(2.2, 2.2))
            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            
            try:
                extracted_text = pytesseract.image_to_string(img, lang="deu+eng")
            except Exception:
                extracted_text = pytesseract.image_to_string(img, lang="eng")
                
            doc.close()
            
            if extracted_text and len(extracted_text.strip()) > 15:
                ai_results = ask_local_ai_to_clean(extracted_text)
                data["Normnummer"] = ai_results["Normnummer"]
                data["Ausgabedatum"] = ai_results["Ausgabedatum"]
                data["Titel_DE"] = ai_results["Titel_DE"]
                
        if data["Normnummer"] == "Nicht gefunden" or len(data["Normnummer"]) < 4:
            clean_name = filename.replace("_", " ").replace("-", " ")
            fn_norm = re.search(r'((?:DIN|NF|MAN|FD|ISO|EN|NFA)\s*[A-Z\d_-]+)', clean_name, re.IGNORECASE)
            if fn_norm: 
                data["Normnummer"] = fn_norm.group(1).strip().upper()
            fn_date = re.search(r'(\b\d{4}\s*\d{2}\b|\b\d{4}-\d{2}\b|\b\d{2}\.\d{4}\b)', clean_name)
            if fn_date: 
                data["Ausgabedatum"] = fn_date.group(1).strip()

    except Exception as e:
        print(f"Fehler bei {filename}: {str(e)}")
        
    data["Dateiname"] = remove_illegal_excel_characters(data["Dateiname"])
    data["Normnummer"] = remove_illegal_excel_characters(data["Normnummer"])
    data["Ausgabedatum"] = remove_illegal_excel_characters(data["Ausgabedatum"])
    data["Titel_DE"] = remove_illegal_excel_characters(data["Titel_DE"])
    return data

# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_to_excel(results: list, output_path: str):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normen Übersicht"
    
    headers = ["Dateiname", "Normnummer", "Ausgabedatum", "Titel (DE)", "Status"]
    ws.append(headers)
    
    header_fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
    for row_data in results:
        ws.append([row_data["Dateiname"], row_data["Normnummer"], row_data["Ausgabedatum"], row_data["Titel_DE"], row_data["Status"]])
        
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)

# ---------------------------------------------------------------------------
# Grafische Benutzeroberfläche (Tkinter GUI)
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Universal KI Normen-Extraktor V14 (by Noel Joan)")
        self.geometry("600x350")
        self.input_dir = tk.StringVar()
        self.output_file = tk.StringVar()
        self.create_widgets()
        
    def create_widgets(self):
        frame1 = ttk.LabelFrame(self, text="1. PDF-Quellordner auswählen", padding=10)
        frame1.pack(fill="x", padx=10, pady=5)
        ttk.Entry(frame1, textvariable=self.input_dir, width=50).pack(side="left", padx=5)
        ttk.Button(frame1, text="Durchsuchen", command=self.browse_input).pack(side="left")
        
        frame2 = ttk.LabelFrame(self, text="2. Excel-Zieldatei auswählen", padding=10)
        frame2.pack(fill="x", padx=10, pady=5)
        ttk.Entry(frame2, textvariable=self.output_file, width=50).pack(side="left", padx=5)
        ttk.Button(frame2, text="Speichern unter", command=self.browse_output).pack(side="left")
        
        self.btn_start = ttk.Button(self, text="KI-Extraktion starten", command=self.start_processing)
        self.btn_start.pack(pady=15)
        
        self.progress = ttk.Progressbar(self, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(pady=5)
        self.lbl_status = ttk.Label(self, text="Bereit. Lokale KI 'Qwen' ist verbunden.")
        self.lbl_status.pack()

    def browse_input(self):
        path = filedialog.askdirectory()
        if path: self.input_dir.set(path)
        
    def browse_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel-Dateien", "*.xlsx")])
        if path: self.output_file.set(path)

    def start_processing(self):
        if not self.input_dir.get() or not self.output_file.get():
            messagebox.showerror("Fehler", "Bitte beide Pfade angeben!")
            return
        self.btn_start.config(state="disabled")
        threading.Thread(target=self.process_files, daemon=True).start()
    def process_files(self):
        try:
            pdf_folder = Path(self.input_dir.get())
            pdf_files = list(pdf_folder.glob("*.pdf"))
            
            if not pdf_files:
                self.lbl_status.config(text="Keine PDFs gefunden.")
                return
                
            self.progress["maximum"] = len(pdf_files)
            results = []
            
            for idx, pdf_path in enumerate(pdf_files, 1):
                self.lbl_status.config(text=f"KI analysiert: {pdf_path.name} ({idx}/{len(pdf_files)})")
                self.progress["value"] = idx
                self.update_idletasks()
                
                data = extract_metadata_from_pdf_content(str(pdf_path))
                results.append(data)
                
            export_to_excel(results, self.output_file.get())
            self.lbl_status.config(text="Fertig!")
            messagebox.showinfo("Erfolg", "KI-Extraktion erfolgreich abgeschlossen!")
            
        except Exception as crash_err:
            messagebox.showerror("Kritischer Verarbeitungsfehler", f"Das Skript musste abbrechen:\n\n{str(crash_err)}")
        finally:
            self.btn_start.config(state="normal")

App().mainloop()
